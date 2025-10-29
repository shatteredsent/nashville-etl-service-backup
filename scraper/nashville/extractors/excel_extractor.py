"""
Excel document extractor.
Handles .xlsx and .xls files with multiple sheets.
"""
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from .base_extractor import BaseExtractor


class ExcelExtractor(BaseExtractor):
    """Extract structured data from Excel files."""

    SUPPORTED_EXTENSIONS = {'.xlsx', '.xlsm', '.xls'}

    # Common column name variations (same as CSV)
    NAME_COLUMNS = frozenset(
        ['name', 'event', 'venue', 'title', 'event_name', 'venue_name'])
    ADDRESS_COLUMNS = frozenset(
        ['address', 'venue_address', 'location', 'street'])
    DATE_COLUMNS = frozenset(
        ['date', 'event_date', 'start_date', 'datetime', 'when'])
    DESCRIPTION_COLUMNS = frozenset(
        ['description', 'desc', 'details', 'info', 'notes'])
    URL_COLUMNS = frozenset(['url', 'website', 'link', 'web', 'webpage'])
    PHONE_COLUMNS = frozenset(
        ['phone', 'telephone', 'contact', 'phone_number'])
    CITY_COLUMNS = frozenset(['city', 'venue_city', 'location_city'])

    def extract(self) -> List[Dict[str, Any]]:
        """Extract structured data from Excel file."""
        try:
            workbook = self._load_workbook()
            all_items = []

            for sheet in workbook.worksheets:
                self._log(f"Processing sheet: {sheet.title}")
                items = self._process_sheet(sheet)
                all_items.extend(items)

            workbook.close()

            self._log(f"Extracted {len(all_items)} total items from Excel")
            return [self._create_item(**item) for item in all_items if self._is_valid(item)]

        except Exception as e:
            self._log(f"Excel extraction failed: {e}", 'error')
            raise

    def _load_workbook(self):
        """Load Excel workbook."""
        try:
            return load_workbook(str(self.file_path), data_only=True)
        except Exception as e:
            raise RuntimeError(f"Failed to load Excel file: {e}")

    def _process_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Process a single worksheet."""
        rows = list(sheet.iter_rows(values_only=True))

        if not rows or len(rows) < 2:
            self._log(
                f"Sheet '{sheet.title}' has insufficient data", 'warning')
            return []

        header = self._extract_header(rows[0])
        if not header:
            self._log(f"No valid header in sheet '{sheet.title}'", 'warning')
            return []

        column_mapping = self._map_columns(header)
        if not column_mapping:
            self._log(
                f"No recognizable columns in sheet '{sheet.title}'", 'warning')
            return []

        return self._parse_rows(rows[1:], column_mapping)

    def _extract_header(self, header_row: tuple) -> List[str]:
        """Extract and normalize header row."""
        return [
            str(cell).strip().lower()
            for cell in header_row
            if cell is not None
        ]

    def _map_columns(self, header: List[str]) -> Dict[str, int]:
        """Map Excel columns to standardized field names."""
        mapping = {}

        for idx, col in enumerate(header):
            if col in self.NAME_COLUMNS:
                mapping['name'] = idx
            elif col in self.ADDRESS_COLUMNS:
                mapping['venue_address'] = idx
            elif col in self.DATE_COLUMNS:
                mapping['event_date'] = idx
            elif col in self.DESCRIPTION_COLUMNS:
                mapping['description'] = idx
            elif col in self.URL_COLUMNS:
                mapping['url'] = idx
            elif col in self.PHONE_COLUMNS:
                mapping['phone'] = idx
            elif col in self.CITY_COLUMNS:
                mapping['venue_city'] = idx

        return mapping

    def _parse_rows(
        self,
        rows: List[tuple],
        column_mapping: Dict[str, int]
    ) -> List[Dict[str, Any]]:
        """Parse data rows into structured items."""
        items = []

        for row in rows:
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            item = {}
            for field, col_idx in column_mapping.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    value = str(row[col_idx]).strip()
                    if value and value.lower() != 'none':
                        item[field] = value

            if item.get('name'):
                item['venue_name'] = item['name']
                items.append(item)

        return items

    def _is_valid(self, item: Dict) -> bool:
        """Validate item has required fields."""
        return bool(item.get('name') and len(item['name']) >= 2)
